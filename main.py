import xlsxwriter
import os
import sys
if hasattr(sys, 'frozen'):
    os.environ['PATH'] = sys._MEIPASS + ";" + os.environ['PATH']
from openpyxl import load_workbook
from mainUI import Ui_MainWindow
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtCore import QTimer
from time import sleep

from splashScreen import Ui_SplashScreen

## GLOBALS
counter = 0

class App():
    def readFile(self,file):
        if file.endswith(".xlsx"):
            try:
                self.wbRef = load_workbook(file)
            except:
                print("Please import valid source file.")
        else:
            print("This file format is not supported. (.xlsx)")

    def Search(self,path):
        if self.wbRef:
            if path:
                try:
                    wsRef = self.wbRef.get_sheet_by_name(self.wbRef.sheetnames[0])
                    self.soldImgNames = []
                    self.unsoldcontent = []
                    self.content, self.soldImgNames = self.loadSheetContent(self,wsRef)
                    self.findImage(self,path)
                    self.bIsSearchComplete = True
                    print("Search complete.")
                except:
                    print("Search ERROR.")
            else:
                print("Please select a images path.")
        else:
            print("Please select a source path.")
        
    def loadSheetContent(self,ws):
        rows = ws.rows
        columns = ws.columns
        content, imgNames = [], []
        for row in rows:
            line = []
            imgNames.append(row[0].value)
            for col in row:
                line.append(col.value)
            content.append(line)
        return content, imgNames
        
    def findImage(self,path):
        for dirPath, dirNames, fileNames in os.walk(path):
            for name in fileNames:
                if name.endswith('.png') or name.endswith('.jpg'):
                    path = os.path.join(path, dirPath, name)
                    base = os.path.splitext(name)[0]
                    bIsVaildName = False
                    for i in range(len(self.soldImgNames)):
                        if base == self.soldImgNames[i]:
                            self.content[i][2] = path
                            bIsVaildName = True
                            
                    if bIsVaildName == False:
                        self.unsoldcontent.append([base,path])
        
    def exportSoldReport(self):
        if self.bIsSearchComplete:
            try:
                self.printProgressBar(self, 0, len(self.content), prefix = 'Progress:', suffix = 'Complete', length = 50)
                wb = xlsxwriter.Workbook('Sold Report.xlsx')
                ws = wb.add_worksheet()
                ws.set_column('B:B', 24)
                ws.set_column('C:C', 18.43)
                for i in range(len(self.content)):
                    index = 0
                    for j in range(len(self.content[i])-1):
                        ws.write(i,j,self.content[i][j])
                        index += 1
                    if self.content[i][index] != None:
                        if self.content[i][index].endswith('.png') or self.content[i][index].endswith('.jpg'):
                            ws.set_row(i,102)
                            ws.insert_image(i,index,self.content[i][index], {'x_scale':0.075, 'y_scale':0.075, 'object_position':3})
                        else:
                            ws.write(i,index,self.content[i][index])
                    else:
                        ws.write(i,index,self.content[i][index])
                    
                    sleep(0.1)
                    self.printProgressBar(self, i + 1, len(self.content), prefix = 'Progress:', suffix = 'Complete', length = 50) 

                wb.close()
                print("Sold report export complete.")
            except:
                print("Sold report export ERROR.")
        else:
            print("Please search images first.")

    def exportUnsoldReport(self):
        if self.bIsSearchComplete:
            try:
                self.printProgressBar(self, 0, len(self.unsoldcontent), prefix = 'Progress:', suffix = 'Complete', length = 50)
                wb = xlsxwriter.Workbook('Unsold Report.xlsx')
                ws = wb.add_worksheet()
                ws.set_column('B:B', 18.43)
                for i in range(len(self.unsoldcontent)):
                    index = 0
                    for j in range(len(self.unsoldcontent[i])-1):
                        ws.write(i,j,self.unsoldcontent[i][j])
                        index += 1
                    if self.unsoldcontent[i][index].endswith('.png') or self.unsoldcontent[i][index].endswith('.jpg'):
                        ws.set_row(i,102)
                        ws.insert_image(i,index,self.unsoldcontent[i][index], {'x_scale':0.075, 'y_scale':0.075, 'object_position':3})

                    sleep(0.1)
                    self.printProgressBar(self, i + 1, len(self.unsoldcontent), prefix = 'Progress:', suffix = 'Complete', length = 50)
                
                wb.close()
                print("Unsold report export complete.")
            except:
                print("Unsold report export ERROR.")
        else:
            print("Please search images first.")
    
    def printProgressBar (self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█'):
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))  
        filledLength = int(length * iteration // total)  
        bar = fill * filledLength + '-' * (length - filledLength)  
        print('\r%s |%s| %s%% %s' % (prefix, bar, percent, suffix), end = '\r')  
        # Print New Line on Complete  
        if iteration == total:   
            print()  
    
class AppWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

    def eventExcelBrowseData(self):
        try:
            data = QtWidgets.QFileDialog.getOpenFileName(self, 'Open File',os.getcwd(),"All Files(*xlsx)")
            if data:
                self.ui.excellineEdit.setText(data[0])
                App.readFile(App,self.ui.excellineEdit.text())
                App.writeFile(App)
        except:
            pass
    
    def eventImageBrowseData(self):
        try:
            data = QtWidgets.QFileDialog.getExistingDirectory(None, 'Open Dir',os.getcwd())
            newData = ""
            if data:
                for c in data:
                    if c == "/":
                        c = "\\"
                    newData += c
                self.ui.imagelineEdit.setText(newData)
        except:
            pass

    def eventSearch(self):
        App.Search(App,self.ui.imagelineEdit.text())
    
    def eventExportSoldReport(self):
        App.exportSoldReport(App)

    def eventExportUnsoldReport(self):
        App.exportUnsoldReport(App)

class splashWindow(QMainWindow, Ui_SplashScreen):
    def __init__(self):
        super().__init__()
        self.ui = Ui_SplashScreen()
        self.ui.setupUi(self)
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        self.timer = QTimer(self)
        self.timer.start(35)
        self.timer.timeout.connect(self.progress)
        self.show()

    def progress(self):
        global counter
        self.ui.progressBar.setValue(counter)

        if counter > 100:
            self.timer.stop()
            self.main = AppWindow()
            self.main.show()
            self.close()
        
        counter += 2


if __name__ == "__main__":
    App.wbRef = None
    App.bIsSearchComplete = False
    app = QApplication(sys.argv)
    win = splashWindow()
    sys.exit(app.exec_())